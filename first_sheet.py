import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
wb=load_workbook('Deleted.xlsx')
sheet=wb.sheetnames
sheet = wb[sheet[0]]
#Initialize variable
count_all=0
count_newp=0
count_p=0
count_row=0
count_col=0
count_new_reply=0
count_positive=0
count_neutral=0
count_negitave=0
new_list = []
old_list = []

#Counting the rows and columns and create a array
for row in sheet.rows:
	count_col+=1
	for cell in row:
		if count_col ==1:
			count_row+=1
print(count_row,"*",count_col)
arr = [[0 for i in range(0,count_row)] for j in range(0,count_col)]

#讀取全部的資料，存入Array
for row in sheet.rows:
	for cell in row:
		cell_str=str(cell.value)
		cell_col=int(ord(cell.column)-65)
		cell_row=int(cell.row-1)
		arr[cell_row][cell_col]=cell_str
		#print("{:>4s}".format(cell.coordinate),end=" ")#表格欄位
	#print(end="\n")

#計算新文章、正面、負面數，新文章標題存入new_list
for i in range(0,cell_row+1):
	for j in range(0,cell_col+1):
		if j == 5:
			if arr[i][j] == "主文":
				count_newp+=1
				new_list.append(arr[i][j-4])
				count_all+=1
		if j == 12:
			if arr[i][j] == "正面":
				count_positive+=1
			elif arr[i][j] == "中立":
				count_neutral+=1
			elif arr[i][j] == "負面":
				count_negitave+=1

#計算總文章數和回文數，若標題不存在於new_list或old_list就存入old_list
for i in range(0,cell_row+1):
	for j in range(0,cell_col+1):
		if j ==1:
			if not (i==0)|(arr[i][j]==' '):
				count_p+=1
			if not (i==0):
				if not(arr[i][j] in new_list)|(arr[i][j] in old_list):
					count_all+=1
					if not arr[i][j+4] == "主文":
						old_list.append(arr[i][j])

#計算新文章的回文引發數
for i in range(0,cell_row+1):
	for j in range(0,cell_col+1):
		if j == 5:
			if (arr[i][j-4]in new_list)&("回文" in arr[i][j]):
				count_new_reply+=1

print("文章數：",count_p,end="\n")
print("主題數：",count_all,end="\n")
print("新文章：",count_newp,end="\n")
print("新文章引發回文數：",count_new_reply,end="\n")
print("總計：",count_newp+count_new_reply,end="\n")
print("回文數：",count_p-count_newp,end="\n")
print("舊文引發回文：",count_p-count_newp-count_new_reply,end="\n")
print("新主題內容比例：",float((count_newp+count_new_reply)/count_p),end="\n")
print("新主題比例：",float(count_newp/count_all),end="\n")
print("正面：",count_positive,end="\n")
print("中立：",count_neutral,end="\n")
print("負面：",count_negitave,end="\n")
print("PN：",float(count_positive/count_negitave),end="\n")